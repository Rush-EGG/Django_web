from django.shortcuts import render, redirect, HttpResponse
from django.core.files.uploadedfile import InMemoryUploadedFile
from openpyxl import load_workbook

from app01 import models

from app01.utils.pagination import Pagination


def depart_list(request):
    # 部门列表

    # 去数据库中找到所有的部门信息
    queryset = models.Department.objects.all()

    query_object = Pagination(request, queryset)
    context = {
        'queryset': query_object.page_queryset,
        'page_string': query_object.html()
    }

    return render(request, "depart_list.html", context)


def depart_add(request):
    # 添加部门
    if request.method == "GET":
        return render(request, "depart_add.html")

    # 获取用户提交的数据
    title = request.POST.get("title")

    # 保存到数据库
    models.Department.objects.create(title=title)

    # 回到显示页面
    return redirect("/depart/list")


def depart_delete(request):
    # 删除部门

    # 获取id
    nid = request.GET.get('nid')
    # 删除
    models.Department.objects.filter(id=nid).delete()

    # 回到显示页面
    return redirect("/depart/list")
    # return HttpResponse("删除成功")


def depart_edit(request, nid):
    # 编辑部门

    if request.method == "GET":
        # 根据nid获取该行的数据
        row_object = models.Department.objects.filter(id=nid).first()
        # print(row_object.id, row_object.title)

        return render(request, "depart_edit.html", {"row_object": row_object})

    # 得到用户提交的标题
    title = request.POST.get("title")
    # 根据id进行修改
    models.Department.objects.filter(id=nid).update(title=title)

    return redirect("/depart/list")


def depart_multi(request):
    # 批量上传excel文件

    # 获取用户上传的文件对象
    file_object = request.FILES.get('exc')
    # print(type(file_object)) # <class 'django.core.files.uploadedfile.InMemoryUploadedFile'>

    # 把对象传递给openpyxl,读取文件的内容
    wb = load_workbook(file_object)
    # 得到该excel的第一个sheet
    sheet = wb.worksheets[0]

    # 得到第一行第二列单元格的内容
    # cell = sheet.cell(1, 2)
    # print(cell.value)

    # 循环获取从第二行开始每一行的数据
    for row in sheet.iter_rows(min_row=2):
        content = row[0].value
        # print(content)
        exists = models.Department.objects.filter(title=content).exists()
        if not exists:
            models.Department.objects.create(title=content)

    return redirect('/depart/list/')
